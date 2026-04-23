"""
Build equity research charts workbook — snapshot analysis across 8 residential REITs.

Runs end-to-end:
  1. Auto-detects the latest Saturday-ending scrape week
  2. Loads that week's raw CSVs, applying the same known-issue fixes as build_excel.py
     (AMH deposit-offer bare-percent false positive, ESS/UDR week-1 scraper bugs)
  3. Generates 14 publication-ready charts as native Excel bar charts
  4. Renders the two hero charts (Unit Concession Rate + NER Discount) as
     high-resolution PNGs via matplotlib, embeds them alongside the native charts
  5. Saves output/REIT_Research_Charts_<WEEK_ENDING>.xlsx

Called automatically each Saturday by weekly_run.py after the Excel rebuild.
"""

import re
from datetime import datetime, date, timedelta
from pathlib import Path

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = Path(__file__).resolve().parent
RAW_DIR = BASE_DIR / "data" / "raw"
OUT_DIR = BASE_DIR / "output"
OUT_DIR.mkdir(exist_ok=True)

# ── Style constants ───────────────────────────────────────────────────────
NAVY = "#003A70"
DARK_BLUE = "#1a4e8a"
MID_BLUE = "#4472C4"
LIGHT_BLUE = "#6b9bd2"
STEEL = "#8FAADC"
SLATE = "#A9C4E0"
ICE = "#C5D9F1"
PALE = "#DCE6F1"
GOLD = "#D4A843"
RED = "#C00000"
DARK_RED = "#8B0000"
GREEN = "#548235"
GRAY = "#808080"
LIGHT_GRAY = "#D9D9D9"
BG_COLOR = "#FAFBFD"
GRID_COLOR = "#E8ECF1"
TEXT_COLOR = "#2C3E50"
SUBTLE_TEXT = "#7F8C9B"
BAR_GRADIENT = [PALE, ICE, SLATE, STEEL, LIGHT_BLUE, MID_BLUE, DARK_BLUE, NAVY]

# Excel fill constants (openpyxl wants no '#')
X_NAVY = NAVY.lstrip("#")
X_MID = MID_BLUE.lstrip("#")
X_LIGHT = LIGHT_BLUE.lstrip("#")
X_GOLD = GOLD.lstrip("#")
X_RED = RED.lstrip("#")
X_GREEN = GREEN.lstrip("#")
X_GRAY = GRAY.lstrip("#")
X_LIGHT_GRAY = LIGHT_GRAY.lstrip("#")

HEADER_FILL = PatternFill("solid", fgColor=X_NAVY)
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
DATA_FONT = Font(name="Arial", size=9)
TITLE_FONT = Font(name="Arial", bold=True, size=11, color=X_NAVY)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


# ── Week detection + data load ────────────────────────────────────────────

_DATE_RE = re.compile(r"_raw_(\d{4}-\d{2}-\d{2})")


def detect_latest_week() -> tuple[date, list[Path]]:
    """
    Scan data/raw/ and return (week_ending_saturday, files_in_that_week).
    Files whose date is Sat–Fri (inclusive) belong to the same week bucket.
    """
    files = sorted(RAW_DIR.glob("*_raw_*.csv"))
    buckets: dict[date, list[Path]] = {}
    for f in files:
        m = _DATE_RE.search(f.stem)
        if not m:
            continue
        d = datetime.strptime(m.group(1), "%Y-%m-%d").date()
        # Anchor to Saturday ON OR BEFORE (matches build_excel.py convention).
        # Python weekday: Mon=0..Sat=5..Sun=6. Saturday-on-or-before = d - (d.weekday()-5)%7.
        wk_anchor = d - timedelta(days=(d.weekday() - 5) % 7)
        buckets.setdefault(wk_anchor, []).append(f)
    if not buckets:
        raise RuntimeError("No scrape files found in data/raw/")
    latest = max(buckets)
    return latest, buckets[latest]


def load_clean_week(files: list[Path]) -> pd.DataFrame:
    """Load CSVs, dedupe on (reit, unit_id), apply known-issue fixes."""
    df = pd.concat([pd.read_csv(f) for f in files], ignore_index=True)
    df["scrape_date"] = pd.to_datetime(df["scrape_date"])
    df = df.sort_values("scrape_date").groupby(["reit", "unit_id"]).last().reset_index()

    # AMH fix: null bare-percent concession_raw (security deposit offer FP).
    bare_mask = (
        (df["reit"] == "AMH")
        & df["concession_raw"].fillna("").str.match(r"^\s*\d+\s*%\s*off\s*$", case=False)
    )
    conc_cols = ["has_concession", "concession_hardness", "concession_raw",
                 "concession_type", "concession_value", "concession_pct_lease_value",
                 "concession_pct_lease_term", "effective_monthly_rent"]
    if bare_mask.any():
        df["has_concession"] = df["has_concession"].astype("object")
        for col in conc_cols:
            if col in df.columns:
                df.loc[bare_mask, col] = None
        df["has_concession"] = df["has_concession"].fillna(False).astype(bool)
        print(f"  [FIX] Nulled {int(bare_mask.sum()):,} AMH bare-percent rows.")

    df["beds_clean"] = pd.to_numeric(df["beds"], errors="coerce")
    df["rent_psf"] = df["rent"] / df["sqft"]
    return df


# ── Excel styling helpers ─────────────────────────────────────────────────

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def style_data(ws, start_row, end_row, max_col, pct_cols=None):
    pct_cols = pct_cols or []
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col in pct_cols:
                cell.number_format = "0.0%"
            elif isinstance(cell.value, float):
                if cell.value > 100:
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "0.00"


def write_sorted_bar(ws, title_text, chart_title, data_series, sort_asc=True,
                     color=X_NAVY, val_fmt="#,##0", axis_fmt=None,
                     extra_cols=None, pct_cols=None, horizontal=True,
                     chart_anchor="D3", chart_w=22, chart_h=14):
    ws.cell(1, 1, title_text).font = TITLE_FONT
    if sort_asc:
        data_series = data_series.sort_values(ascending=True)

    ws.cell(3, 1, data_series.index.name or "REIT")
    ws.cell(3, 2, data_series.name or "Value")
    n_extra = 0
    if extra_cols:
        for j, name in enumerate(extra_cols.keys()):
            ws.cell(3, 3 + j, name)
            n_extra += 1

    for i, (label, val) in enumerate(data_series.items()):
        ws.cell(4 + i, 1, label)
        ws.cell(4 + i, 2, val)
        if extra_cols:
            for j, series in enumerate(extra_cols.values()):
                ws.cell(4 + i, 3 + j, series.get(label, None))

    total_cols = 2 + n_extra
    style_header(ws, 3, total_cols)
    style_data(ws, 4, 3 + len(data_series), total_cols, pct_cols=pct_cols or [])

    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    chart.title = chart_title
    chart.style = 2
    chart.width = chart_w
    chart.height = chart_h
    chart.legend = None
    data_ref = Reference(ws, min_col=2, min_row=3, max_row=3 + len(data_series))
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=3 + len(data_series))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    s = chart.series[0]
    s.graphicalProperties.solidFill = color
    if axis_fmt:
        chart.y_axis.numFmt = axis_fmt
    s.dLbls = DataLabelList()
    s.dLbls.showVal = True
    s.dLbls.numFmt = val_fmt
    ws.add_chart(chart, chart_anchor)
    return chart


# ── Matplotlib helpers ────────────────────────────────────────────────────

def setup_mpl():
    plt.rcParams.update({
        "font.family": "Arial", "font.size": 11,
        "axes.facecolor": BG_COLOR, "figure.facecolor": "white",
        "axes.edgecolor": GRID_COLOR, "axes.linewidth": 0.5,
        "xtick.color": TEXT_COLOR, "ytick.color": TEXT_COLOR,
        "text.color": TEXT_COLOR, "axes.labelcolor": TEXT_COLOR,
        "figure.dpi": 200,
    })


def render_chart1_png(df: pd.DataFrame, wk: date, out_path: Path):
    """Hero chart 1: % of listed units offering concessions."""
    setup_mpl()
    conc = df.groupby("reit")["has_concession"].mean().sort_values() * 100
    reits, values = conc.index.tolist(), conc.values
    fig, ax = plt.subplots(figsize=(10, 6.5))
    n = len(values)
    colors = [BAR_GRADIENT[int(i / max(n - 1, 1) * (len(BAR_GRADIENT) - 1))] for i in range(n)]
    bars = ax.barh(reits, values, height=0.62, color=colors, edgecolor="white", linewidth=0.5, zorder=3)
    for bar, val in zip(bars, values):
        if val > 5:
            ax.text(bar.get_width() - 1.5, bar.get_y() + bar.get_height() / 2,
                    f"{val:.1f}%", va="center", ha="right",
                    fontsize=12, fontweight="bold", color="white")
        else:
            ax.text(bar.get_width() + 1.2, bar.get_y() + bar.get_height() / 2,
                    f"{val:.1f}%", va="center", ha="left",
                    fontsize=12, fontweight="bold", color=NAVY)
    ax.set_xlim(0, 100)
    ax.xaxis.set_major_formatter(mtick.PercentFormatter())
    ax.xaxis.set_major_locator(mtick.MultipleLocator(20))
    ax.tick_params(axis="y", labelsize=13, length=0)
    ax.tick_params(axis="x", labelsize=10, length=0)
    ax.set_axisbelow(True)
    ax.xaxis.grid(True, color=GRID_COLOR, linewidth=0.7)
    ax.yaxis.grid(False)
    for spine in ["top", "right", "bottom"]:
        ax.spines[spine].set_visible(False)
    ax.spines["left"].set_color(GRID_COLOR)
    fig.text(0.06, 0.96, "% of Listed Units Offering Concessions",
             fontsize=16, fontweight="bold", color=NAVY, ha="left")
    fig.text(0.06, 0.915,
             f"Snapshot across 8 publicly traded residential REITs  |  Week ending {wk}",
             fontsize=10, color=SUBTLE_TEXT, ha="left")
    fig.text(0.06, 0.02,
             "Source: REIT websites, scraped unit-level data. AMH and INVH are single-family rental REITs "
             "with no rent concessions (deposit-offer false positive filtered).",
             fontsize=7.5, color=SUBTLE_TEXT, ha="left", style="italic")
    plt.tight_layout(rect=[0.0, 0.05, 1.0, 0.90])
    fig.savefig(str(out_path), dpi=200, bbox_inches="tight", facecolor="white")
    plt.close()


def render_chart3_png(df: pd.DataFrame, wk: date, out_path: Path):
    """Hero chart 3: NER discount depth."""
    setup_mpl()
    c = df[df["has_concession"] & df["effective_monthly_rent"].notna()].copy()
    c["disc"] = (1 - c["effective_monthly_rent"] / c["rent"]) * 100
    disc = c.groupby("reit")["disc"].mean().sort_values()
    reits, values = disc.index.tolist(), disc.values
    fig, ax = plt.subplots(figsize=(10, 5.5))
    RED_GRAD = ["#FADBD8", "#F1948A", "#E74C3C", "#CB4335", "#B03A2E", "#922B21"]
    n = len(values)
    colors = [RED_GRAD[int(i / max(n - 1, 1) * (len(RED_GRAD) - 1))] for i in range(n)]
    bars = ax.barh(reits, values, height=0.58, color=colors, edgecolor="white", linewidth=0.5, zorder=3)
    avg_rents = c.groupby("reit")["rent"].mean()
    avg_ner = c.groupby("reit")["effective_monthly_rent"].mean()
    max_v = values.max() if len(values) else 0
    for bar, reit, val in zip(bars, reits, values):
        ax.text(bar.get_width() - 0.15, bar.get_y() + bar.get_height() / 2,
                f"{val:.1f}%", va="center", ha="right",
                fontsize=12, fontweight="bold",
                color="white" if val > 5 else TEXT_COLOR)
        gap = avg_rents.get(reit, 0) - avg_ner.get(reit, 0)
        ax.text(bar.get_width() + 0.2, bar.get_y() + bar.get_height() / 2,
                f"${gap:,.0f}/mo avg gap",
                va="center", ha="left", fontsize=8.5, color=SUBTLE_TEXT)
    ax.set_xlim(0, max(max_v * 1.25, 12))
    ax.xaxis.set_major_formatter(mtick.PercentFormatter())
    ax.xaxis.set_major_locator(mtick.MultipleLocator(2))
    ax.tick_params(axis="y", labelsize=13, length=0)
    ax.tick_params(axis="x", labelsize=10, length=0)
    ax.set_axisbelow(True)
    ax.xaxis.grid(True, color=GRID_COLOR, linewidth=0.7)
    ax.yaxis.grid(False)
    for spine in ["top", "right", "bottom"]:
        ax.spines[spine].set_visible(False)
    ax.spines["left"].set_color(GRID_COLOR)
    fig.text(0.06, 0.96, "Avg Discount to Asking Rent for Concession Units",
             fontsize=16, fontweight="bold", color=DARK_RED, ha="left")
    fig.text(0.06, 0.91,
             f"Net Effective Rent vs Gross Asking Rent  |  Concession-offering units only  |  Week ending {wk}",
             fontsize=10, color=SUBTLE_TEXT, ha="left")
    fig.text(0.06, 0.02,
             "Source: REIT websites, scraped unit-level data. Excludes REITs with no concessions listed. "
             "Discount = 1 - (NER / Gross Rent).",
             fontsize=7.5, color=SUBTLE_TEXT, ha="left", style="italic")
    plt.tight_layout(rect=[0.0, 0.05, 1.0, 0.89])
    fig.savefig(str(out_path), dpi=200, bbox_inches="tight", facecolor="white")
    plt.close()


# ── Main workbook build ───────────────────────────────────────────────────

def build_workbook(df: pd.DataFrame, wk: date) -> Path:
    mf = df[df["reit"].isin(["AVB", "CPT", "EQR", "ESS", "MAA", "UDR"])]

    wb = Workbook()

    # 1. Unit concession rate
    ws1 = wb.active
    ws1.title = "Unit Concession Rate"
    conc_rate = df.groupby("reit")["has_concession"].mean()
    conc_rate.name = "% Units w/ Concession"
    write_sorted_bar(ws1, f"Chart 1: % of Listed Units Offering Concessions by REIT (Week ending {wk})",
                     "% of Listed Units Offering Concessions", conc_rate,
                     color=X_NAVY, val_fmt="0.0%", axis_fmt="0%", pct_cols=[2])

    # 2. Community concession rate
    ws2 = wb.create_sheet("Community Conc Rate")
    comm = df.groupby(["reit", "community"]).agg(cr=("has_concession", "mean")).reset_index()
    comm_rate = comm.groupby("reit").apply(lambda x: (x["cr"] > 0).mean())
    comm_rate.name = "% Communities w/ Concession"
    write_sorted_bar(ws2, f"Chart 2: % of Communities Offering Any Concession (Week ending {wk})",
                     "% of Communities Offering Any Concession", comm_rate,
                     color=X_GOLD, val_fmt="0.0%", axis_fmt="0%", pct_cols=[2])

    # 3. NER discount
    ws3 = wb.create_sheet("NER Discount")
    c = df[df["has_concession"] & df["effective_monthly_rent"].notna()].copy()
    c["discount"] = 1 - c["effective_monthly_rent"] / c["rent"]
    disc = c.groupby("reit")["discount"].mean()
    disc.name = "Avg Discount to Asking"
    write_sorted_bar(ws3, f"Chart 3: Avg Discount to Asking Rent (Concession Units, Week ending {wk})",
                     "Avg NER Discount to Asking Rent", disc,
                     color=X_RED, val_fmt="0.0%", axis_fmt="0.0%", pct_cols=[2])

    # 4. Concession type mix (stacked)
    ws4 = wb.create_sheet("Conc Type Mix")
    ws4.cell(1, 1, f"Chart 4: Concession Type Mix by REIT (Week ending {wk})").font = TITLE_FONT
    conc_units = df[df["has_concession"]]
    type_mix = conc_units.groupby(["reit", "concession_type"]).size().unstack(fill_value=0)
    type_pct = type_mix.div(type_mix.sum(axis=1), axis=0)
    type_cols = [c for c in ["months_free", "dollar_off", "percent_off", "date_free"] if c in type_pct.columns]
    type_pct = type_pct[type_cols]
    ws4.cell(3, 1, "REIT")
    for j, col in enumerate(type_cols):
        ws4.cell(3, 2 + j, col.replace("_", " ").title())
    for i, reit in enumerate(type_pct.index):
        ws4.cell(4 + i, 1, reit)
        for j, col in enumerate(type_cols):
            ws4.cell(4 + i, 2 + j, type_pct.loc[reit, col])
    style_header(ws4, 3, 1 + len(type_cols))
    style_data(ws4, 4, 3 + len(type_pct), 1 + len(type_cols),
               pct_cols=list(range(2, 2 + len(type_cols))))
    chart4 = BarChart()
    chart4.type = "col"
    chart4.grouping = "stacked"
    chart4.title = "Concession Type Mix by REIT"
    chart4.style = 2
    chart4.width, chart4.height = 22, 14
    data4 = Reference(ws4, min_col=2, max_col=1 + len(type_cols), min_row=3, max_row=3 + len(type_pct))
    cats4 = Reference(ws4, min_col=1, min_row=4, max_row=3 + len(type_pct))
    chart4.add_data(data4, titles_from_data=True)
    chart4.set_categories(cats4)
    chart4.y_axis.numFmt = "0%"
    for idx, s in enumerate(chart4.series):
        s.graphicalProperties.solidFill = [X_NAVY, X_GOLD, X_GREEN, X_GRAY][idx % 4]
    ws4.add_chart(chart4, "D3")

    # 5. Market concession rate
    ws5 = wb.create_sheet("Market Conc Rate")
    ws5.cell(1, 1, f"Chart 5: Concession Rate by Market (100+ units, Week ending {wk})").font = TITLE_FONT
    mkt = df.groupby("market").agg(
        units=("unit_id", "count"),
        conc_rate=("has_concession", "mean"),
        reits=("reit", "nunique"),
    ).reset_index()
    mkt = mkt[mkt["units"] >= 100].sort_values("conc_rate", ascending=True).tail(20)
    ws5.cell(3, 1, "Market")
    ws5.cell(3, 2, "Concession Rate")
    ws5.cell(3, 3, "Units")
    ws5.cell(3, 4, "# REITs")
    for i, (_, row) in enumerate(mkt.iterrows()):
        ws5.cell(4 + i, 1, row["market"])
        ws5.cell(4 + i, 2, row["conc_rate"])
        ws5.cell(4 + i, 3, row["units"])
        ws5.cell(4 + i, 4, row["reits"])
    style_header(ws5, 3, 4)
    style_data(ws5, 4, 3 + len(mkt), 4, pct_cols=[2])
    chart5 = BarChart()
    chart5.type = "bar"
    chart5.title = "Concession Rate by Market"
    chart5.style = 2
    chart5.width, chart5.height = 24, 16
    chart5.legend = None
    data5 = Reference(ws5, min_col=2, min_row=3, max_row=3 + len(mkt))
    cats5 = Reference(ws5, min_col=1, min_row=4, max_row=3 + len(mkt))
    chart5.add_data(data5, titles_from_data=True)
    chart5.set_categories(cats5)
    chart5.series[0].graphicalProperties.solidFill = X_LIGHT
    chart5.y_axis.numFmt = "0%"
    chart5.series[0].dLbls = DataLabelList()
    chart5.series[0].dLbls.showVal = True
    chart5.series[0].dLbls.numFmt = "0%"
    ws5.add_chart(chart5, "F3")

    # 6. Avg asking rent
    ws6 = wb.create_sheet("Avg Asking Rent")
    avg_rent = df.groupby("reit")["rent"].mean()
    avg_rent.name = "Avg Asking Rent"
    write_sorted_bar(ws6, f"Chart 6: Average Asking Rent by REIT (Week ending {wk})",
                     "Average Asking Rent ($)", avg_rent,
                     color=X_GREEN, val_fmt="$#,##0", axis_fmt="$#,##0")

    # 7. Rent PSF
    ws7 = wb.create_sheet("Rent PSF")
    psf = df.groupby("reit")["rent_psf"].mean()
    psf.name = "Avg Rent PSF"
    write_sorted_bar(ws7, f"Chart 7: Average Asking Rent PSF by REIT (Week ending {wk})",
                     "Average Asking Rent PSF ($/sqft)", psf,
                     color=X_NAVY, val_fmt="$0.00", axis_fmt="$0.00")

    # 8. Market rent PSF
    ws8 = wb.create_sheet("Market Rent PSF")
    top_mkts = df.groupby("market").size().sort_values(ascending=False).head(15).index
    mkt_psf = df[df["market"].isin(top_mkts)].groupby("market")["rent_psf"].mean()
    mkt_psf.name = "Avg Rent PSF"
    mkt_units = df[df["market"].isin(top_mkts)].groupby("market").size()
    write_sorted_bar(ws8, f"Chart 8: Avg Rent PSF by Market (Top 15 by Unit Count, Week ending {wk})",
                     "Avg Rent PSF by Market", mkt_psf,
                     color=X_GOLD, val_fmt="$0.00", axis_fmt="$0.00",
                     extra_cols={"Units": mkt_units}, chart_w=24, chart_h=16, chart_anchor="E3")

    # 9. Concession rate by beds
    ws9 = wb.create_sheet("Conc Rate by Beds")
    ws9.cell(1, 1, f"Chart 9: Concession Rate by REIT and Bedroom Count (MF, Week ending {wk})").font = TITLE_FONT
    mf_beds = mf[mf["beds_clean"].isin([0, 1, 2, 3])].copy()
    pvt = mf_beds.groupby(["reit", "beds_clean"])["has_concession"].mean().unstack()
    pvt.columns = ["Studio", "1BR", "2BR", "3BR"]
    ws9.cell(3, 1, "REIT")
    for j, col in enumerate(pvt.columns):
        ws9.cell(3, 2 + j, col)
    for i, reit in enumerate(pvt.index):
        ws9.cell(4 + i, 1, reit)
        for j, col in enumerate(pvt.columns):
            val = pvt.loc[reit, col]
            ws9.cell(4 + i, 2 + j, val if not pd.isna(val) else None)
    style_header(ws9, 3, 5)
    style_data(ws9, 4, 3 + len(pvt), 5, pct_cols=[2, 3, 4, 5])
    for i in range(len(pvt)):
        for j in range(4):
            cell = ws9.cell(4 + i, 2 + j)
            if cell.value is not None:
                v = float(cell.value)
                if v >= 0.75:
                    cell.fill = PatternFill("solid", fgColor="F4CCCC")
                elif v >= 0.50:
                    cell.fill = PatternFill("solid", fgColor="FCE5CD")
                elif v >= 0.25:
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                else:
                    cell.fill = PatternFill("solid", fgColor="D9EAD3")
    chart9 = BarChart()
    chart9.type = "col"
    chart9.grouping = "clustered"
    chart9.title = "Concession Rate by REIT and Bed Count"
    chart9.style = 2
    chart9.width, chart9.height = 22, 14
    data9 = Reference(ws9, min_col=2, max_col=5, min_row=3, max_row=3 + len(pvt))
    cats9 = Reference(ws9, min_col=1, min_row=4, max_row=3 + len(pvt))
    chart9.add_data(data9, titles_from_data=True)
    chart9.set_categories(cats9)
    chart9.y_axis.numFmt = "0%"
    for idx, s in enumerate(chart9.series):
        s.graphicalProperties.solidFill = [X_LIGHT_GRAY, X_NAVY, X_GOLD, X_GREEN][idx]
    ws9.add_chart(chart9, "G3")

    # 10. Sample size
    ws10 = wb.create_sheet("Sample Size")
    units = df.groupby("reit")["unit_id"].count()
    units.name = "# Units Listed"
    communities = df.groupby("reit")["community"].nunique()
    write_sorted_bar(ws10, f"Chart 10: Available Units Listed by REIT (Week ending {wk})",
                     "Available Units Listed by REIT", units,
                     color=X_NAVY, val_fmt="#,##0",
                     extra_cols={"# Communities": communities}, chart_anchor="E3")

    # 11. Gross vs NER
    ws11 = wb.create_sheet("Gross vs NER")
    ws11.cell(1, 1, f"Chart 11: Gross Asking vs Net Effective Rent (Concession Units, Week ending {wk})").font = TITLE_FONT
    c = df[df["has_concession"] & df["effective_monthly_rent"].notna()].copy()
    gross = c.groupby("reit")["rent"].mean()
    ner = c.groupby("reit")["effective_monthly_rent"].mean()
    paired = pd.DataFrame({"Gross Rent": gross, "Net Effective": ner}).sort_values("Gross Rent")
    ws11.cell(3, 1, "REIT")
    ws11.cell(3, 2, "Gross Asking Rent")
    ws11.cell(3, 3, "Net Effective Rent")
    ws11.cell(3, 4, "Discount")
    for i, reit in enumerate(paired.index):
        ws11.cell(4 + i, 1, reit)
        ws11.cell(4 + i, 2, paired.loc[reit, "Gross Rent"])
        ws11.cell(4 + i, 3, paired.loc[reit, "Net Effective"])
        ws11.cell(4 + i, 4, 1 - paired.loc[reit, "Net Effective"] / paired.loc[reit, "Gross Rent"])
    style_header(ws11, 3, 4)
    style_data(ws11, 4, 3 + len(paired), 4, pct_cols=[4])
    chart11 = BarChart()
    chart11.type = "col"
    chart11.grouping = "clustered"
    chart11.title = "Gross vs Net Effective Rent"
    chart11.style = 2
    chart11.width, chart11.height = 22, 14
    data11 = Reference(ws11, min_col=2, max_col=3, min_row=3, max_row=3 + len(paired))
    cats11 = Reference(ws11, min_col=1, min_row=4, max_row=3 + len(paired))
    chart11.add_data(data11, titles_from_data=True)
    chart11.set_categories(cats11)
    chart11.y_axis.numFmt = "$#,##0"
    chart11.series[0].graphicalProperties.solidFill = X_NAVY
    chart11.series[1].graphicalProperties.solidFill = X_RED
    ws11.add_chart(chart11, "F3")

    # 12. Rent distribution (percentiles)
    ws12 = wb.create_sheet("Rent Distribution")
    ws12.cell(1, 1, f"Chart 12: Asking Rent Distribution by REIT (Week ending {wk})").font = TITLE_FONT
    pctiles = df.groupby("reit")["rent"].quantile([0.10, 0.25, 0.50, 0.75, 0.90]).unstack()
    pctiles.columns = ["P10", "P25", "Median", "P75", "P90"]
    pctiles = pctiles.sort_values("Median")
    ws12.cell(3, 1, "REIT")
    for j, col in enumerate(pctiles.columns):
        ws12.cell(3, 2 + j, col)
    for i, reit in enumerate(pctiles.index):
        ws12.cell(4 + i, 1, reit)
        for j, col in enumerate(pctiles.columns):
            ws12.cell(4 + i, 2 + j, pctiles.loc[reit, col])
    style_header(ws12, 3, 6)
    style_data(ws12, 4, 3 + len(pctiles), 6)
    chart12 = BarChart()
    chart12.type = "col"
    chart12.grouping = "clustered"
    chart12.title = "Asking Rent Distribution"
    chart12.style = 2
    chart12.width, chart12.height = 24, 14
    data12 = Reference(ws12, min_col=2, max_col=6, min_row=3, max_row=3 + len(pctiles))
    cats12 = Reference(ws12, min_col=1, min_row=4, max_row=3 + len(pctiles))
    chart12.add_data(data12, titles_from_data=True)
    chart12.set_categories(cats12)
    chart12.y_axis.numFmt = "$#,##0"
    for idx, s in enumerate(chart12.series):
        s.graphicalProperties.solidFill = [X_LIGHT_GRAY, X_LIGHT, X_NAVY, X_GOLD, X_GRAY][idx]
    ws12.add_chart(chart12, "H3")

    # 13. Avg unit size
    ws13 = wb.create_sheet("Avg Unit Size")
    sqft = df.groupby("reit")["sqft"].mean()
    sqft.name = "Avg Sqft"
    write_sorted_bar(ws13, f"Chart 13: Average Unit Size by REIT (sqft, Week ending {wk})",
                     "Average Unit Size (sqft)", sqft,
                     color=X_GREEN, val_fmt="#,##0")

    # 14. Concession intensity (breadth x depth)
    ws14 = wb.create_sheet("Conc Intensity")
    ws14.cell(1, 1, f"Chart 14: Concession Intensity = Breadth x Depth (MF, Week ending {wk})").font = TITLE_FONT
    breadth = mf.groupby("reit")["has_concession"].mean()
    c_mf = mf[mf["has_concession"] & mf["effective_monthly_rent"].notna()].copy()
    c_mf["disc"] = 1 - c_mf["effective_monthly_rent"] / c_mf["rent"]
    depth = c_mf.groupby("reit")["disc"].mean()
    intensity = pd.DataFrame({
        "Breadth (% w/ Conc)": breadth,
        "Depth (Avg Discount)": depth,
    }).dropna()
    intensity["Intensity (B x D)"] = intensity["Breadth (% w/ Conc)"] * intensity["Depth (Avg Discount)"]
    intensity = intensity.sort_values("Intensity (B x D)", ascending=True)
    ws14.cell(3, 1, "REIT")
    ws14.cell(3, 2, "Breadth (% w/ Conc)")
    ws14.cell(3, 3, "Depth (Avg Discount)")
    ws14.cell(3, 4, "Intensity (B x D)")
    for i, reit in enumerate(intensity.index):
        ws14.cell(4 + i, 1, reit)
        ws14.cell(4 + i, 2, intensity.loc[reit, "Breadth (% w/ Conc)"])
        ws14.cell(4 + i, 3, intensity.loc[reit, "Depth (Avg Discount)"])
        ws14.cell(4 + i, 4, intensity.loc[reit, "Intensity (B x D)"])
    style_header(ws14, 3, 4)
    style_data(ws14, 4, 3 + len(intensity), 4, pct_cols=[2, 3, 4])
    chart14 = BarChart()
    chart14.type = "col"
    chart14.grouping = "clustered"
    chart14.title = "Concession Intensity: Breadth vs Depth"
    chart14.style = 2
    chart14.width, chart14.height = 22, 14
    data14 = Reference(ws14, min_col=2, max_col=3, min_row=3, max_row=3 + len(intensity))
    cats14 = Reference(ws14, min_col=1, min_row=4, max_row=3 + len(intensity))
    chart14.add_data(data14, titles_from_data=True)
    chart14.set_categories(cats14)
    chart14.y_axis.numFmt = "0%"
    chart14.series[0].graphicalProperties.solidFill = X_NAVY
    chart14.series[1].graphicalProperties.solidFill = X_RED
    ws14.add_chart(chart14, "F3")

    # ── Render + embed hero PNGs ─────────────────────────────────────────
    png1 = OUT_DIR / f"chart1_unit_concession_rate_{wk}.png"
    png3 = OUT_DIR / f"chart3_ner_discount_{wk}.png"
    render_chart1_png(df, wk, png1)
    render_chart3_png(df, wk, png3)

    img1 = Image(str(png1))
    img1.width, img1.height = 900, 585
    ws1.add_image(img1, "D18")
    img3 = Image(str(png3))
    img3.width, img3.height = 900, 495
    ws3.add_image(img3, "D16")

    out_path = OUT_DIR / f"REIT_Research_Charts_{wk}.xlsx"
    wb.save(str(out_path))
    return out_path


def main():
    print("[Charts] Detecting latest scrape week...")
    wk, files = detect_latest_week()
    print(f"[Charts] Latest week: {wk}  ({len(files)} CSV files)")
    print("[Charts] Loading + cleaning data...")
    df = load_clean_week(files)
    print(f"[Charts] {len(df):,} rows after dedupe and fixes.")
    print("[Charts] Building workbook...")
    out = build_workbook(df, wk)
    size_kb = out.stat().st_size / 1024
    print(f"[Charts] Saved: {out}  ({size_kb:.0f} KB)")


if __name__ == "__main__":
    main()
